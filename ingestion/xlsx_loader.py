import openpyxl
from loguru import logger


def load_xlsx(file_path: str) -> str:
    # read_only=True streams rows without loading the entire workbook into RAM.
    # This is the critical fix: previously load_workbook() loaded everything
    # eagerly, making large XLSX files very slow to process.
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    text = ""

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        text += f"Sheet: {sheet_name}\n"

        for row in sheet.iter_rows(values_only=True):
            row_text = " | ".join(
                str(cell) for cell in row if cell is not None
            )
            if row_text.strip():
                text += row_text + "\n"

    workbook.close()
    logger.info(f"XLSX loaded: {file_path}")
    return text